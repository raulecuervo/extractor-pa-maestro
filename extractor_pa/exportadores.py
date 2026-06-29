# -*- coding: utf-8 -*-
"""
Adaptadores de salida (Fase 6).

Convierten el modelo canónico (`ResultadoExtraccion`) a formatos usables por
otros sistemas, sin acoplar el extractor a ninguno:

- **JSON**  → `exportar_json` (stdlib).
- **CSV**   → `exportar_csv` (stdlib; una tabla por archivo).
- **Excel** → `exportar_excel` (openpyxl; una tabla por hoja).
- **DataFrame** → `a_dataframes` (pandas, dependencia opcional).

Y la versión **consolidada multi-plan** (`*_consolidado`), que apila los
indicadores/alertas/financiero de varias políticas en tablas únicas (con
columnas `politica` y `archivo`), para analizar todo el corpus junto.

Las tablas se entregan en **formato ancho**: las metas anuales se expanden a
columnas `meta_<año>` (la unión de años de todos los planes incluidos).
"""

from __future__ import annotations

import csv
import json
import os
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# ─────────────────────────── aplanado ───────────────────────────

def _flat_indicador(ind, meta, tipo: str) -> dict:
    """Aplana un IR/IP a un dict plano: metadatos + campos + meta_<año>."""
    d = asdict(ind)
    metas = d.pop("metas_por_anio", {}) or {}
    fila = {
        "politica": meta.nombre_politica or "",
        "archivo": meta.archivo_fuente or "",
        "formato": meta.formato_detectado or "",
        "tipo": tipo,
    }
    fila.update(d)
    for anio in sorted(metas):
        fila[f"meta_{anio}"] = metas[anio]
    return fila


def _flat_alerta(a, meta) -> dict:
    d = asdict(a)
    d["archivo"] = meta.archivo_fuente or d.get("archivo_fuente", "")
    return d


def _flat_financiero(f, meta) -> dict:
    d = asdict(f)
    d["politica"] = meta.nombre_politica or ""
    d["archivo"] = meta.archivo_fuente or ""
    return d


def tablas(resultado) -> dict:
    """Devuelve {nombre_tabla: [filas_planas]} para un resultado."""
    m = resultado.metadatos
    meta_fila = dict(asdict(m))
    meta_fila["anios_detectados"] = ", ".join(str(a) for a in (m.anios_detectados or []))
    return {
        "metadatos": [meta_fila],
        "indicadores_resultado": [_flat_indicador(i, m, "IR")
                                  for i in resultado.indicadores_resultado],
        "indicadores_producto": [_flat_indicador(i, m, "IP")
                                 for i in resultado.indicadores_producto],
        "alertas": [_flat_alerta(a, m) for a in resultado.alertas],
        "financiero": [_flat_financiero(f, m) for f in resultado.financiero],
    }


def tablas_consolidadas(resultados: Iterable) -> dict:
    """Apila las tablas de varios resultados en tablas únicas (multi-plan)."""
    out = {"metadatos": [], "indicadores_resultado": [],
           "indicadores_producto": [], "alertas": [], "financiero": []}
    for res in resultados:
        for nombre, filas in tablas(res).items():
            out[nombre].extend(filas)
    return out


# ─────────────────────────── helpers de escritura ───────────────

def _columnas(filas: list) -> list:
    """Unión ordenada de columnas (primer-visto) a través de todas las filas."""
    cols = {}
    for f in filas:
        for k in f:
            cols.setdefault(k, None)
    return list(cols)


def _celda(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        v = json.dumps(v, ensure_ascii=False, default=str)
    if isinstance(v, str):
        # Elimina caracteres de control ilegales para Excel (0x00-0x1F salvo
        # tab/salto), que rompen openpyxl y aparecen en textos cualitativos.
        return ILLEGAL_CHARACTERS_RE.sub("", v)
    return v


# ─────────────────────────── JSON ───────────────────────────────

def exportar_json(resultado, ruta: str | Path) -> str:
    """Escribe el resultado completo como JSON (estructura canónica anidada)."""
    with open(ruta, "w", encoding="utf-8") as fh:
        json.dump(resultado.to_dict(), fh, ensure_ascii=False, indent=2, default=str)
    return str(ruta)


def exportar_json_consolidado(resultados: Iterable, ruta: str | Path) -> str:
    with open(ruta, "w", encoding="utf-8") as fh:
        json.dump([r.to_dict() for r in resultados], fh,
                  ensure_ascii=False, indent=2, default=str)
    return str(ruta)


# ─────────────────────────── CSV ────────────────────────────────

def _escribir_csv_tablas(tbls: dict, carpeta: str | Path, prefijo: str = "") -> list:
    os.makedirs(carpeta, exist_ok=True)
    escritos = []
    for nombre, filas in tbls.items():
        if not filas:
            continue
        cols = _columnas(filas)
        ruta = os.path.join(str(carpeta), f"{prefijo}{nombre}.csv")
        with open(ruta, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for f in filas:
                w.writerow({c: _celda(f.get(c)) for c in cols})
        escritos.append(ruta)
    return escritos


def exportar_csv(resultado, carpeta: str | Path, prefijo: str = "") -> list:
    """Escribe un CSV por tabla (metadatos, IR, IP, alertas, financiero)."""
    return _escribir_csv_tablas(tablas(resultado), carpeta, prefijo)


def exportar_csv_consolidado(resultados: Iterable, carpeta: str | Path,
                             prefijo: str = "consolidado_") -> list:
    return _escribir_csv_tablas(tablas_consolidadas(resultados), carpeta, prefijo)


# ─────────────────────────── Excel ──────────────────────────────

def _escribir_excel_tablas(tbls: dict, ruta: str | Path) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, filas in tbls.items():
        ws = wb.create_sheet(title=nombre[:31])
        if not filas:
            ws.append(["(sin datos)"])
            continue
        cols = _columnas(filas)
        ws.append(cols)
        for f in filas:
            ws.append([_celda(f.get(c)) for c in cols])
    if not wb.sheetnames:
        wb.create_sheet(title="vacio")
    wb.save(ruta)
    return str(ruta)


def exportar_excel(resultado, ruta: str | Path) -> str:
    """Escribe un .xlsx con una hoja por tabla."""
    return _escribir_excel_tablas(tablas(resultado), ruta)


def exportar_excel_consolidado(resultados: Iterable, ruta: str | Path) -> str:
    return _escribir_excel_tablas(tablas_consolidadas(resultados), ruta)


# ─────────────────────────── DataFrame (pandas opcional) ────────

def a_dataframes(resultado) -> dict:
    """Devuelve {nombre_tabla: DataFrame}. Requiere pandas (extra 'pandas')."""
    try:
        import pandas as pd
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "a_dataframes requiere pandas. Instala con: pip install extractor-pa[pandas]"
        ) from e
    return {n: pd.DataFrame(filas) for n, filas in tablas(resultado).items()}


def a_dataframes_consolidado(resultados: Iterable) -> dict:
    try:
        import pandas as pd
    except ImportError as e:  # pragma: no cover
        raise ImportError("a_dataframes_consolidado requiere pandas.") from e
    return {n: pd.DataFrame(filas) for n, filas in tablas_consolidadas(resultados).items()}
