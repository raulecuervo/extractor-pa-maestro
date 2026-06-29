# -*- coding: utf-8 -*-
"""
Smoke test de la Fase 1: construye un Excel sintético con el FORMATO NUEVO y
verifica la extracción de punta a punta:
  - detección de formato y hoja,
  - resolución de columnas por encabezado,
  - forward-fill (celdas combinadas: la 2ª fila del IR va vacía),
  - deduplicación de IR,
  - escala % en las metas (celdas con number_format '0%').

Ejecutable como script (`python tests/test_smoke.py`) o vía pytest.
"""

from __future__ import annotations

import os
import sys
import tempfile

from openpyxl import Workbook

# Permite importar el paquete sin instalarlo.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from extractor_pa import extraer_plan_accion, NIVEL_ERROR  # noqa: E402
from extractor_pa.vigencia import calcular_vigencia  # noqa: E402


def _nuevo_ws():
    """Crea un workbook con la hoja y los encabezados del formato nuevo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de acción"
    _escribir_encabezados(ws)
    return wb, ws


def _escribir_encabezados(ws) -> None:
    """Escribe metadatos y encabezados (filas 9/10/11) del formato nuevo."""
    # --- Metadatos de cabecera ---
    ws.cell(row=4, column=2, value="Política Pública de Prueba")
    ws.cell(row=7, column=2, value="Sector Salud")
    ws.cell(row=7, column=8, value="Secretaría de Salud")
    ws.cell(row=8, column=1, value="Objetivo general de prueba")
    ws.cell(row=3, column=1, value="Documento CONPES 9999")

    # --- Encabezados fila 10 (anclas principales del formato nuevo) ---
    h10 = {
        3: "Resultado esperado",
        4: "Nombre del indicador de resultado",
        5: "Vigente/No Vigente",
        6: "Importancia relativa del resultado (%)",
        7: "Fórmula del indicador de resultado",
        8: "Sector Responsable",
        9: "Entidad Responsable",
        10: "ODS",
        11: "Meta ODS",
        12: "Tipo de anualización",
        13: "Periodicidad",
        21: "Meta de resultado Final",   # exacto (lo usa el detector)
        22: "Producto esperado",         # exacto (lo usa el detector)
    }
    for c, v in h10.items():
        ws.cell(row=10, column=c, value=v)

    # --- Encabezados fila 9 (grupos) ---
    ws.cell(row=9, column=41, value="Meta de producto Final")
    ws.cell(row=9, column=42, value="Responsables de la ejecución")
    ws.cell(row=9, column=45, value="Corresponsables de la ejecución")

    # --- Encabezados fila 11 (repetidos + metas anuales) ---
    h11 = {
        14: "Valor", 15: "Año", 16: "Fuente",
        17: "Fecha de inicio", 18: "Fecha de finalización",
        19: "Meta 2024", 20: "Meta 2025",
        34: "Valor", 35: "Año", 36: "Fuente",
        37: "Fecha de inicio", 38: "Fecha de finalización",
        39: "Meta 2024", 40: "Meta 2025",
    }
    for c, v in h11.items():
        ws.cell(row=11, column=c, value=v)


def _construir_excel(ruta: str) -> None:
    """Crea un plan de acción mínimo en formato nuevo (caso base Fase 1)."""
    wb, ws = _nuevo_ws()

    # --- Fila 12: IR 1.1 + IP 1.1.1 ---
    fila12 = {
        1: "1. Objetivo uno", 2: 60, 3: "1.1 Resultado uno", 4: "IR uno",
        5: "Vigente", 6: 100, 7: "(a/b)*100", 8: "Salud",
        9: "Secretaría de Salud", 10: "3", 11: "3.4", 12: "Creciente",
        13: "Anual", 14: 10, 15: 2019, 16: "DANE", 17: "2024-01-01",
        18: "2030-12-31", 19: 12, 20: 20, 21: 20,
        22: "1.1.1 Producto uno", 23: "IP uno", 24: "Vigente", 25: 50,
        26: "(c/d)", 27: "Suma", 28: "Trimestral", 29: "OBJ PDD",
        30: "Meta PDD", 31: "PROY-1", 32: "Género", 33: None,
        34: 0, 35: 2019, 36: "SDP", 37: "2024-01-01", 38: "2030-12-31",
        42: "Salud", 43: "Secretaría de Salud", 44: "Dir 1",
        45: "Educación", 46: "Sec Educación", 47: "Dir 2",
    }
    for c, v in fila12.items():
        ws.cell(row=12, column=c, value=v)
    # Metas IP como porcentaje (number_format '0%'): 0.05 -> 5%, 0.10 -> 10%
    for c, v in ((39, 0.05), (40, 0.10), (41, 0.10)):
        cell = ws.cell(row=12, column=c, value=v)
        cell.number_format = "0%"

    # --- Fila 13: 2º IP del MISMO IR (columnas IR vacías → prueba forward-fill) ---
    fila13 = {22: "1.1.2 Producto dos", 23: "IP dos", 24: "Vigente", 25: 50}
    for c, v in fila13.items():
        ws.cell(row=13, column=c, value=v)
    for c, v in ((39, 0.07), (40, 0.13), (41, 0.13)):
        cell = ws.cell(row=13, column=c, value=v)
        cell.number_format = "0%"

    # --- Fila 14: IR 2.1 + IP 2.1.1 (metas en número plano, sin %) ---
    fila14 = {
        1: "2. Objetivo dos", 2: 40, 3: "2.1 Resultado dos", 4: "IR dos",
        5: "Vigente", 6: 100, 12: "Constante", 13: "Anual",
        19: 5, 20: 8, 21: 8,
        22: "2.1.1 Producto tres", 23: "IP tres", 24: "Vigente", 25: 100,
        39: 3, 40: 5, 41: 5,
    }
    for c, v in fila14.items():
        ws.cell(row=14, column=c, value=v)

    wb.save(ruta)


def test_extraccion_formato_nuevo():
    ruta = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    _construir_excel(ruta)

    res = extraer_plan_accion(ruta)

    # No debe haber ERRORES de extracción.
    errores = [a for a in res.alertas if a.nivel == NIVEL_ERROR]
    assert not errores, f"Errores inesperados: {[a.descripcion for a in errores]}"

    # Metadatos y formato.
    assert res.metadatos.nombre_politica == "Política Pública de Prueba"
    assert res.metadatos.formato_detectado == "nuevo"
    assert res.metadatos.hoja_usada == "Plan de acción"
    assert res.metadatos.documento_conpes and "CONPES" in res.metadatos.documento_conpes
    assert res.metadatos.anios_detectados == [2024, 2025]

    # 2 IR (dedup correcto pese a que 1.1 aparece en 2 filas) y 3 IP.
    assert len(res.indicadores_resultado) == 2, [i.codigo_ir for i in res.indicadores_resultado]
    assert len(res.indicadores_producto) == 3, [i.codigo_ip for i in res.indicadores_producto]

    ir = {i.codigo_ir: i for i in res.indicadores_resultado}
    assert set(ir) == {"1.1", "2.1"}
    assert ir["1.1"].codigo_objetivo == "1"
    assert ir["1.1"].nombre_indicador == "IR uno"
    assert ir["1.1"].metas_por_anio == {2024: 12.0, 2025: 20.0} or \
           ir["1.1"].metas_por_anio == {2024: 12, 2025: 20}
    assert ir["1.1"].meta_final in (20, 20.0)
    assert ir["1.1"].escala_pct is False  # metas IR en número plano

    ip = {i.codigo_ip: i for i in res.indicadores_producto}
    assert set(ip) == {"1.1.1", "1.1.2", "2.1.1"}

    # El 2º IP heredó su IR padre vía forward-fill.
    assert ip["1.1.2"].codigo_ir == "1.1"
    assert ip["1.1.2"].codigo_objetivo == "1"

    # Escala %: 0.05 -> 5.0, 0.10 -> 10.0; meta_final 0.10 -> 10.0
    assert ip["1.1.1"].escala_pct is True
    assert ip["1.1.1"].metas_por_anio == {2024: 5.0, 2025: 10.0}
    assert ip["1.1.1"].meta_final == 10.0

    # IP del IR 2 con metas planas (sin %).
    assert ip["2.1.1"].escala_pct is False
    assert ip["2.1.1"].metas_por_anio == {2024: 3.0, 2025: 5.0}


def _construir_excel_fase2(ruta: str) -> None:
    """Plan que ejercita la Fase 2:
    - IR 1.1: 1ª fila 'No Vigente' (peso 0) y 2ª fila 'Vigente' (peso 80)
      -> la ascensión debe promover la versión vigente.
    - IR 2.1: con 'Sector Responsable' VACÍO -> el ffill agrupado NO debe
      heredar el sector del IR 1.1 (no contaminación)."""
    wb, ws = _nuevo_ws()

    # IR 1.1 — fila histórica No Vigente (se leerá primero).
    f12 = {
        1: "1. Objetivo uno", 2: 100, 3: "1.1 Resultado uno",
        4: "IR viejo", 5: "No Vigente", 6: 0, 8: "Salud (histórico)",
        9: "Sec Salud", 12: "Creciente", 13: "Anual", 21: 9, 19: 8, 20: 9,
        22: "1.1.1 Producto uno", 23: "IP uno", 24: "No Vigente", 25: 0,
        39: 5, 40: 10,
    }
    for c, v in f12.items():
        ws.cell(row=12, column=c, value=v)

    # IR 1.1 — fila VIGENTE actual (columna 'resultado' vacía: misma agrupación).
    f13 = {
        4: "IR nuevo", 5: "Vigente", 6: 80, 8: "Salud", 9: "Sec Salud",
        12: "Creciente", 13: "Anual", 21: 10,
        22: "1.1.2 Producto dos", 23: "IP dos", 24: "Vigente", 25: 100,
        39: 5, 40: 10,
    }
    for c, v in f13.items():
        ws.cell(row=13, column=c, value=v)

    # IR 2.1 — Sector Responsable (col 8) AUSENTE a propósito.
    f14 = {
        1: "2. Objetivo dos", 2: 0, 3: "2.1 Resultado dos",
        4: "IR dos", 5: "Vigente", 6: 20, 9: "Sec X",
        12: "Constante", 13: "Anual", 21: 5, 19: 3, 20: 5,
        22: "2.1.1 Producto tres", 23: "IP tres", 24: "Vigente", 25: 100,
        39: 2, 40: 3,
    }
    for c, v in f14.items():
        ws.cell(row=14, column=c, value=v)

    wb.save(ruta)


def test_ascension_fila_vigente():
    """La versión VIGENTE (fila inferior) debe ganar a la histórica (fila superior)."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_fase2_extractor.xlsx")
    _construir_excel_fase2(ruta)
    res = extraer_plan_accion(ruta)

    assert not [a for a in res.alertas if a.nivel == NIVEL_ERROR]
    ir = {i.codigo_ir: i for i in res.indicadores_resultado}
    assert "1.1" in ir
    # Campos de identidad promovidos desde la fila vigente:
    assert ir["1.1"].nombre_indicador == "IR nuevo"
    assert ir["1.1"].es_vigente == "Vigente"
    assert float(ir["1.1"].peso_pct) == 80.0


def test_no_contaminacion_entre_ir():
    """Un IR con Sector vacío NO debe heredar el Sector del IR anterior."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_fase2_extractor.xlsx")
    _construir_excel_fase2(ruta)
    res = extraer_plan_accion(ruta)

    ir = {i.codigo_ir: i for i in res.indicadores_resultado}
    assert "2.1" in ir
    # El sector del IR 2.1 debe quedar vacío, NO "Salud" del IR 1.1.
    assert ir["2.1"].sector_responsable is None
    assert ir["2.1"].nombre_indicador == "IR dos"


def test_calcular_vigencia():
    """Lógica pura del año de vigencia (Fase 3)."""
    metas = {2024: 10, 2025: 20, 2026: 30}
    # Año explícito con meta -> ese año; anterior = inmediatamente previo.
    assert calcular_vigencia(metas, 2025) == (2025, 2024, 20, 10)
    # Año explícito SIN meta (futuro) -> año <= más cercano (2026).
    assert calcular_vigencia(metas, 2030) == (2026, 2025, 30, 20)
    # Sin metas -> todo None.
    assert calcular_vigencia({}, 2025) == (None, None, None, None)


def test_vigencia_en_extraccion():
    """El año de vigencia se propaga a IR e IP durante la extracción."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    _construir_excel(ruta)
    res = extraer_plan_accion(ruta, anio_vigencia=2025)

    ir = {i.codigo_ir: i for i in res.indicadores_resultado}
    assert ir["1.1"].anio_vigencia == 2025
    assert ir["1.1"].anio_vigencia_anterior == 2024
    assert ir["1.1"].meta_vigencia_actual in (20, 20.0)
    assert ir["1.1"].meta_vigencia_anterior in (12, 12.0)

    ip = {i.codigo_ip: i for i in res.indicadores_producto}
    # Metas del IP en escala % (0.10 -> 10.0, 0.05 -> 5.0)
    assert ip["1.1.1"].anio_vigencia == 2025
    assert ip["1.1.1"].meta_vigencia_actual == 10.0
    assert ip["1.1.1"].meta_vigencia_anterior == 5.0


def _escribir_ficha(wb, titulo, metodologia, descripcion, fuentes, dias,
                    observaciones, unidad):
    """Crea una hoja de ficha técnica con rótulos en col 1 y valores en col 2."""
    ws = wb.create_sheet(title=titulo)
    ws.cell(row=1, column=1, value="Metodología de medición")
    ws.cell(row=1, column=2, value=metodologia)
    ws.cell(row=2, column=1, value="Descripción")
    ws.cell(row=2, column=2, value=descripcion)
    ws.cell(row=3, column=1, value="Fuentes de información")
    ws.cell(row=3, column=2, value=fuentes)
    ws.cell(row=4, column=1, value="Días de rezago")
    ws.cell(row=4, column=2, value=dias)
    ws.cell(row=5, column=1, value="Observaciones")
    ws.cell(row=5, column=2, value=observaciones)
    # Cuadrícula de unidad de medida: etiqueta en B6, casilla marcada en C6.
    ws.cell(row=6, column=1, value="Unidad de medida")
    ws.cell(row=6, column=2, value=unidad)
    ws.cell(row=6, column=3, value="x")


def _construir_excel_con_fichas(ruta: str) -> None:
    """Plan mínimo (1 IR + 1 IP) con sus hojas de ficha técnica."""
    wb, ws = _nuevo_ws()
    fila = {
        1: "1. Objetivo uno", 2: 100, 3: "1.1 Resultado uno", 4: "IR uno",
        5: "Vigente", 6: 100, 12: "Creciente", 13: "Anual", 19: 5, 20: 6, 21: 6,
        22: "1.1.1 Producto uno", 23: "IP uno", 24: "Vigente", 25: 100,
        39: 3, 40: 5, 41: 5,
    }
    for c, v in fila.items():
        ws.cell(row=12, column=c, value=v)
    _escribir_ficha(wb, "Ficha técnica IR#1.1", "Promedio ponderado",
                    "Mide la adopción", "DANE", "30 días", "Ninguna", "Porcentaje")
    _escribir_ficha(wb, "Ficha técnica IP#1.1.1", "Conteo directo",
                    "Cuenta productos", "SDP", "15", "Sin observaciones", "Número")
    wb.save(ruta)


def test_lectura_fichas_tecnicas():
    """Las hojas de ficha técnica enriquecen el IR y el IP por código."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_fichas_extractor.xlsx")
    _construir_excel_con_fichas(ruta)
    res = extraer_plan_accion(ruta)

    assert not [a for a in res.alertas if a.nivel == NIVEL_ERROR]

    ir = {i.codigo_ir: i for i in res.indicadores_resultado}["1.1"]
    assert ir.metodologia == "Promedio ponderado"
    assert ir.descripcion == "Mide la adopción"
    assert ir.fuente_datos == "DANE"
    assert ir.dias_rezago == 30
    assert ir.observaciones == "Ninguna"
    assert ir.unidad_medida == "Porcentaje"

    ip = {i.codigo_ip: i for i in res.indicadores_producto}["1.1.1"]
    assert ip.metodologia == "Conteo directo"
    assert ip.dias_rezago == 15
    assert ip.unidad_medida == "Número"

    # Si se desactiva, los campos de ficha quedan vacíos.
    res2 = extraer_plan_accion(ruta, leer_fichas_tecnicas=False)
    ir2 = {i.codigo_ir: i for i in res2.indicadores_resultado}["1.1"]
    assert ir2.metodologia is None and ir2.dias_rezago is None


def _construir_excel_fase5(ruta: str) -> None:
    """Plan que ejercita la Fase 5:
    - IR 1.1 aparece en 2 filas con 'Nombre del indicador' DISTINTO -> inconsistencia.
    - El producto 1.1.1 aparece 2 veces -> código de IP duplicado."""
    wb, ws = _nuevo_ws()
    f12 = {
        1: "1. Objetivo uno", 2: 100, 3: "1.1 Resultado uno", 4: "IR uno",
        5: "Vigente", 6: 100, 12: "Creciente", 13: "Anual", 19: 5, 20: 6, 21: 6,
        22: "1.1.1 Producto uno", 23: "IP uno", 24: "Vigente", 25: 50,
    }
    for c, v in f12.items():
        ws.cell(row=12, column=c, value=v)
    # 2ª fila del MISMO IR: nombre del indicador modificado + producto repetido.
    f13 = {
        3: "1.1 Resultado uno", 4: "IR uno MODIFICADO", 5: "Vigente", 6: 100,
        22: "1.1.1 Producto duplicado", 23: "IP dup", 24: "Vigente", 25: 50,
    }
    for c, v in f13.items():
        ws.cell(row=13, column=c, value=v)
    wb.save(ruta)


def test_consistencia_y_duplicados():
    ruta = os.path.join(tempfile.gettempdir(), "plan_fase5_extractor.xlsx")
    _construir_excel_fase5(ruta)
    res = extraer_plan_accion(ruta)

    # 1 IR (dedup) y 2 IP (ambos 1.1.1).
    assert len(res.indicadores_resultado) == 1
    assert len(res.indicadores_producto) == 2

    tipos = [(a.tipo, a.codigo_ir, a.codigo_ip, a.campo) for a in res.alertas]
    # Inconsistencia en el nombre del IR 1.1.
    assert any(t == "inconsistencia_en_ir" and ir == "1.1" and campo == "nombre_indicador"
               for (t, ir, ip, campo) in tipos), tipos
    # Código de IP duplicado 1.1.1.
    assert any(t == "codigo_ip_duplicado" and ip == "1.1.1"
               for (t, ir, ip, campo) in tipos), tipos


def test_plan_limpio_sin_alertas_consistencia():
    """El plan base (bien formado) no genera inconsistencias ni duplicados."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    _construir_excel(ruta)
    res = extraer_plan_accion(ruta)
    tipos = {a.tipo for a in res.alertas}
    assert "inconsistencia_en_ir" not in tipos
    assert "codigo_ip_duplicado" not in tipos


def test_deteccion_nombres_ficha():
    """La detección del código de ficha tolera todas las convenciones reales."""
    from extractor_pa.lector_fichas import codigo_de_hoja_ficha as cod
    # Convenciones observadas en planes reales:
    assert cod("Ficha técnica IR#1.1") == "1.1"
    assert cod("Ficha técnica IP 1.1.1 ") == "1.1.1"
    assert cod("Ficha técnica IR 2.1") == "2.1"
    assert cod("Ficha de producto 1.1.5") == "1.1.5"
    assert cod("R.1.1") == "1.1"           # Cultos
    assert cod("P.1.1.1") == "1.1.1"       # Cultos
    assert cod("R 1.1 Atención efectiva") == "1.1"      # Salud Mental
    assert cod("P 1.1.1 Acciones Cuidadoras") == "1.1.1"  # Salud Mental
    assert cod("1.1.1. Bibliotecarios formados") == "1.1.1"  # LEO
    assert cod("1.2.Apropiación Social") == "1.2"       # LEO
    assert cod("1.1.10") == "1.1.10"       # Hábitat
    assert cod("IR_1.1") == "1.1"          # Pobreza (guion bajo)
    assert cod("IP_1.1.1 Reclutamiento") == "1.1.1"  # Talento Humano
    assert cod("IR#1.4 Irregualres") == "1.4"        # Talento Humano
    # NO deben confundirse con fichas:
    assert cod("Plan de acción") is None
    assert cod("Desplegables 2") is None
    assert cod("Instructivo ficha técnica") is None
    assert cod("Versiones PA") is None
    assert cod("2019") is None             # un año, sin punto


def test_formato_antiguo_cti():
    """Integración: el formato antiguo (con bloque financiero) se extrae bien.

    Se salta si el archivo real no está disponible (mantiene la suite portable)."""
    import pytest
    ruta = (r'C:\Users\RaulEsteban\Proyectos\sispp-gobierno'
            r'\01_planes_accion\plan_accion_pp_cti_v4-25.xlsx')
    if not os.path.exists(ruta):
        pytest.skip('archivo CTI no disponible en este entorno')

    res = extraer_plan_accion(ruta, anio_vigencia=2026)
    assert res.metadatos.formato_detectado == "antiguo"
    assert not [a for a in res.alertas if a.nivel == NIVEL_ERROR]
    assert len(res.indicadores_resultado) > 0
    assert len(res.indicadores_producto) > 0
    # El bloque financiero debe haberse leído.
    assert len(res.financiero) > 0
    # El IP se resuelve por ancla: el estado de vigencia se lee correctamente.
    ip = {i.codigo_ip: i for i in res.indicadores_producto}
    primer = res.indicadores_producto[0]
    assert primer.es_vigente is not None
    assert primer.nombre_indicador
    # Hay registros financieros con costo y con código IP válido.
    assert any(f.costo_estimado is not None and f.codigo_ip for f in res.financiero)


def _construir_excel_reglas(ruta: str, *, limpio: bool) -> None:
    """Plan para probar las reglas V0–V18.
    - limpio=True: ponderaciones y metas coherentes -> sin alertas de negocio.
    - limpio=False: viola V0 (objetivos≠100), V14 (CRECIENTE meta<LB) y V18
      (IP vigente con peso 0)."""
    wb, ws = _nuevo_ws()
    peso_obj = 100 if limpio else 80
    lb_ir = 10 if limpio else 50
    meta_final_ir = 30
    peso_ip = 100 if limpio else 0
    fila = {
        1: "1. Objetivo uno", 2: peso_obj, 3: "1.1 Resultado uno", 4: "IR uno",
        5: "Vigente", 6: 100, 12: "Creciente", 13: "Anual",
        14: lb_ir, 15: 2023, 17: "2024-01-01", 18: "2025-12-31",
        19: 20, 20: 30, 21: meta_final_ir,
        22: "1.1.1 Producto uno", 23: "IP uno", 24: "Vigente", 25: peso_ip,
        27: "Suma", 28: "Anual", 34: 0, 35: 2023, 37: "2024-01-01", 38: "2025-12-31",
        39: 40, 40: 60, 41: 100,
    }
    for c, v in fila.items():
        ws.cell(row=12, column=c, value=v)
    wb.save(ruta)


def test_reglas_negocio_plan_limpio():
    from extractor_pa import validar_reglas
    ruta = os.path.join(tempfile.gettempdir(), "plan_reglas_ok.xlsx")
    _construir_excel_reglas(ruta, limpio=True)
    res = extraer_plan_accion(ruta)
    alertas = validar_reglas(res)
    graves = [a for a in alertas if a.nivel in ("ERROR", "ADVERTENCIA")]
    assert not graves, [(a.tipo, a.descripcion) for a in graves]


def test_reglas_negocio_violaciones():
    from extractor_pa import validar_reglas
    ruta = os.path.join(tempfile.gettempdir(), "plan_reglas_mal.xlsx")
    _construir_excel_reglas(ruta, limpio=False)
    res = extraer_plan_accion(ruta)
    tipos = {a.tipo for a in validar_reglas(res)}
    assert "ponderacion_objetivos" in tipos, tipos      # V0
    assert "meta_vs_linea_base" in tipos, tipos          # V14
    assert "vigencia_ponderacion" in tipos, tipos        # V18


def test_reglas_via_pipeline_flag():
    """El flag incluir_reglas_negocio agrega las alertas V0–V18 al resultado."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_reglas_mal.xlsx")
    _construir_excel_reglas(ruta, limpio=False)
    res = extraer_plan_accion(ruta, incluir_reglas_negocio=True)
    assert any(a.tipo == "ponderacion_objetivos" for a in res.alertas)


def test_exportadores_un_plan():
    """JSON, CSV, Excel y tablas para un solo plan."""
    import json
    import openpyxl
    from extractor_pa import (
        tablas, exportar_json, exportar_csv, exportar_excel,
    )
    ruta = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    _construir_excel(ruta)
    res = extraer_plan_accion(ruta)

    # tablas(): estructura esperada y metas expandidas a columnas meta_<año>.
    tbls = tablas(res)
    assert set(tbls) == {"metadatos", "indicadores_resultado",
                         "indicadores_producto", "alertas", "financiero"}
    assert len(tbls["indicadores_resultado"]) == 2
    fila_ir = tbls["indicadores_resultado"][0]
    assert "meta_2024" in fila_ir and "meta_2025" in fila_ir
    assert fila_ir["politica"] == "Política Pública de Prueba"

    base = tempfile.mkdtemp()
    # JSON
    rj = exportar_json(res, os.path.join(base, "plan.json"))
    datos = json.load(open(rj, encoding="utf-8"))
    assert datos["metadatos"]["nombre_politica"] == "Política Pública de Prueba"
    assert len(datos["indicadores_producto"]) == 3

    # CSV (una por tabla con datos)
    csvs = exportar_csv(res, base)
    nombres = {os.path.basename(c) for c in csvs}
    assert "indicadores_resultado.csv" in nombres
    assert "indicadores_producto.csv" in nombres

    # Excel (una hoja por tabla, reabrible)
    rx = exportar_excel(res, os.path.join(base, "plan.xlsx"))
    wb = openpyxl.load_workbook(rx)
    assert "indicadores_resultado" in wb.sheetnames
    assert "indicadores_producto" in wb.sheetnames
    # La hoja IP tiene encabezado + 3 filas.
    ws = wb["indicadores_producto"]
    assert ws.max_row == 4


def test_exportador_consolidado_multiplan():
    """El consolidado apila varios planes en tablas únicas."""
    from extractor_pa import tablas_consolidadas
    r1 = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    r2 = os.path.join(tempfile.gettempdir(), "plan_fase2_extractor.xlsx")
    _construir_excel(r1)
    _construir_excel_fase2(r2)
    res1 = extraer_plan_accion(r1)
    res2 = extraer_plan_accion(r2)

    tbls = tablas_consolidadas([res1, res2])
    # IPs de ambos planes apilados.
    assert len(tbls["indicadores_producto"]) == (
        len(res1.indicadores_producto) + len(res2.indicadores_producto))
    # Cada fila trae su archivo de origen.
    archivos = {f["archivo"] for f in tbls["indicadores_producto"]}
    assert len(archivos) == 2


def test_seguimiento_xlsb():
    """Integración Fase S1: extracción de un .xlsb de seguimiento real.

    Se salta si el archivo no está disponible (mantiene la suite portable)."""
    import pytest
    from extractor_pa.seguimiento import extraer_seguimiento
    ruta = (r"C:\Users\RaulEsteban\Proyectos\alertas-seguimientos"
            r"\archivos_base\Seguimiento a Productos PP BTI S1-25.xlsb")
    if not os.path.exists(ruta):
        pytest.skip("archivo .xlsb de seguimiento no disponible en este entorno")
    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        pytest.skip("pyxlsb no instalado")

    res = extraer_seguimiento(ruta)
    assert res.exitoso
    assert res.metadatos.tipo_archivo == "productos"
    assert res.metadatos.periodo == "S1" and res.metadatos.anio_reporte == 2025
    assert len(res.metadatos.anios_detectados) > 0
    assert len(res.indicadores) > 0
    # Algún indicador con código N.N.N y series de avances/metas.
    ind = res.indicadores[0]
    assert ind.codigo and "." in ind.codigo
    assert ind.avances or ind.metas
    # Serializable.
    d = res.to_dict()
    assert "indicadores" in d and "metadatos" in d


def test_consolidar_periodo_unidad():
    """Consolidación por período: SUMA suma trimestres; otros toman el último valor."""
    from extractor_pa.seguimiento import consolidar_periodo
    from extractor_pa.seguimiento import IndicadorSeguimiento

    ind = IndicadorSeguimiento(
        codigo="1.1.1", tipo_anualizacion="Suma",
        avances={"2024_Q1": 10, "2024_Q2": 20},
        acumulados={"2024": 30}, metas={"2024": 30}, meta_final=100,
    )
    c = consolidar_periodo(ind, 2024, "S1")
    assert c["trimestres_encontrados"] == [1, 2]
    assert c["avance_consolidado"] == 30      # SUMA → suma de Q1+Q2
    assert c["suma_trimestres"] == 30

    ind.tipo_anualizacion = "Creciente"
    c2 = consolidar_periodo(ind, 2024, "S1")
    assert c2["avance_consolidado"] == 20      # no-SUMA → último valor (Q2)

    # Período sin datos → None.
    assert consolidar_periodo(ind, 2025, "Anual") is None


def test_seguimiento_cruce_y_consolidacion():
    """Integración S2: cruzar el seguimiento de BTI con su plan y consolidar."""
    import pytest
    from extractor_pa.seguimiento import extraer_seguimiento, cruzar_con_plan, consolidar
    plan_path = (r"C:\Users\RaulEsteban\Proyectos\sispp-gobierno"
                 r"\01_planes_accion\PA_BTI_V4-26_DP.xlsx")
    seg_path = (r"C:\Users\RaulEsteban\Proyectos\alertas-seguimientos"
                r"\archivos_base\Seguimiento a Productos PP BTI S1-25.xlsb")
    if not (os.path.exists(plan_path) and os.path.exists(seg_path)):
        pytest.skip("archivos BTI no disponibles")
    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        pytest.skip("pyxlsb no instalado")

    seg = extraer_seguimiento(seg_path)
    plan = extraer_plan_accion(plan_path)
    rep = cruzar_con_plan(seg, plan)
    assert rep["total"] > 0
    assert rep["asociados"] > 0
    # Al menos un indicador quedó emparejado al plan.
    assert any(i.en_plan and i.tipo_plan in ("IR", "IP") for i in seg.indicadores)
    # tipo_anualizacion se capturó del .xlsb.
    assert any(i.tipo_anualizacion for i in seg.indicadores)
    # Consolidación de un año con datos.
    cons = consolidar(seg, 2024, "Anual")
    assert len(cons) > 0
    assert all("avance_consolidado" in c for c in cons)


def test_consistencia_base_vs_nuevo_unidad():
    """Estabilidad, retroactividad e indicadores nuevo/faltante (caso controlado)."""
    from extractor_pa.seguimiento import validar_consistencia
    from extractor_pa.seguimiento import (
        IndicadorSeguimiento, MetadatosSeguimiento, ResultadoSeguimiento,
    )

    base = ResultadoSeguimiento(
        MetadatosSeguimiento(archivo_fuente="base.xlsb"),
        indicadores=[
            IndicadorSeguimiento(codigo="1.1.1", nombre="Indicador X",
                                 corte="Q4", anio_reporte=2024,
                                 avances={"2024_Q4": 0.2}, acumulados={"2024": 0.2},
                                 metas={"2024": 0.2}, tipo_anualizacion="Constante"),
            IndicadorSeguimiento(codigo="9.9.9", nombre="Solo en base", corte="Q4",
                                 anio_reporte=2024),
        ])
    nuevo = ResultadoSeguimiento(
        MetadatosSeguimiento(archivo_fuente="nuevo.xlsb"),
        indicadores=[
            IndicadorSeguimiento(codigo="1.1.1", nombre="Indicador X CAMBIADO",  # estabilidad
                                 corte="Q4", anio_reporte=2025,
                                 avances={"2024_Q4": 0.9},                       # retroactividad
                                 acumulados={"2024": 0.2}, metas={"2024": 0.2},
                                 tipo_anualizacion="Constante"),
            IndicadorSeguimiento(codigo="8.8.8", nombre="Solo en nuevo", corte="Q4",
                                 anio_reporte=2025),
        ])
    tipos = {a.tipo for a in validar_consistencia(base, nuevo)}
    assert "ERROR_ESTABILIDAD" in tipos
    assert "ERROR_RETROACTIVO" in tipos
    assert "INFO_IND_NUEVO" in tipos      # 8.8.8
    assert "INFO_IND_FALTANTE" in tipos   # 9.9.9


def test_semaforo_clasificacion():
    from extractor_pa.seguimiento import semaforo_de
    assert semaforo_de(0.30) == "ROJO"       # 30%
    assert semaforo_de(0.60) == "AMARILLO"   # 60%
    assert semaforo_de(0.90) == "VERDE"      # 90%
    assert semaforo_de(1.40) == "NARANJA"    # 140%
    assert semaforo_de(None) == "SIN_DATO"


def test_exportadores_seguimiento_real():
    """Integración S4: tablas + JSON/CSV/Excel de un seguimiento real."""
    import json
    import openpyxl
    import pytest
    from extractor_pa.seguimiento import (
        extraer_seguimiento, tablas_seguimiento,
        exportar_json_seguimiento, exportar_csv_seguimiento, exportar_excel_seguimiento,
    )
    seg_path = (r"C:\Users\RaulEsteban\Proyectos\alertas-seguimientos"
                r"\archivos_base\Seguimiento a Productos PP BTI S1-25.xlsb")
    if not os.path.exists(seg_path):
        pytest.skip("archivo .xlsb de seguimiento no disponible")
    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        pytest.skip("pyxlsb no instalado")

    res = extraer_seguimiento(seg_path)
    tbls = tablas_seguimiento(res)
    assert set(tbls) == {"metadatos", "indicadores", "avances_trimestrales",
                         "anual", "cualitativo", "alertas"}
    assert len(tbls["indicadores"]) == len(res.indicadores)
    # formato largo: cada fila de avances tiene anio/trimestre/valor
    if tbls["avances_trimestrales"]:
        fila = tbls["avances_trimestrales"][0]
        assert {"codigo", "anio", "trimestre", "valor"} <= set(fila)

    base = tempfile.mkdtemp()
    rj = exportar_json_seguimiento(res, os.path.join(base, "seg.json"))
    d = json.load(open(rj, encoding="utf-8"))
    assert "indicadores" in d and "metadatos" in d

    csvs = exportar_csv_seguimiento(res, base)
    nombres = {os.path.basename(c) for c in csvs}
    assert "seg_indicadores.csv" in nombres
    assert "seg_avances_trimestrales.csv" in nombres

    rx = exportar_excel_seguimiento(res, os.path.join(base, "seg.xlsx"))
    wb = openpyxl.load_workbook(rx)
    assert "indicadores" in wb.sheetnames and "anual" in wb.sheetnames


def test_consistencia_seguimiento_real():
    """Integración S3: validar el par base/nuevo real de BTI sin errores de ejecución."""
    import pytest
    from extractor_pa.seguimiento import extraer_seguimiento, validar_consistencia
    base_p = (r"C:\Users\RaulEsteban\Proyectos\alertas-seguimientos"
              r"\archivos_base\Seguimiento a Productos PP BTI S1-25.xlsb")
    nuevo_p = (r"C:\Users\RaulEsteban\Proyectos\alertas-seguimientos"
               r"\archivos_nuevos\Seguimiento a Productos PP BTI S2-25.xlsb")
    if not (os.path.exists(base_p) and os.path.exists(nuevo_p)):
        pytest.skip("par BTI base/nuevo no disponible")
    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        pytest.skip("pyxlsb no instalado")
    base = extraer_seguimiento(base_p)
    nuevo = extraer_seguimiento(nuevo_p)
    alertas = validar_consistencia(base, nuevo)
    # Debe ejecutar y producir alertas de calidad del reporte (avance/% etc.).
    assert isinstance(alertas, list)
    assert any(a.tipo.startswith("ADVERTENCIA_") for a in alertas)


def _ejecutar_demo():
    """Construye el Excel, extrae y devuelve el resultado (para uso por consola)."""
    ruta = os.path.join(tempfile.gettempdir(), "plan_prueba_extractor.xlsx")
    _construir_excel(ruta)
    return extraer_plan_accion(ruta)


if __name__ == "__main__":
    test_extraccion_formato_nuevo()
    resultado = _ejecutar_demo()
    print("OK — Smoke test Fase 1 superado.")
    print(f"  Política: {resultado.metadatos.nombre_politica}")
    print(f"  Formato:  {resultado.metadatos.formato_detectado}")
    print(f"  Años:     {resultado.metadatos.anios_detectados}")
    print(f"  IR: {len(resultado.indicadores_resultado)} | "
          f"IP: {len(resultado.indicadores_producto)} | "
          f"Alertas: {len(resultado.alertas)}")
    for ip in resultado.indicadores_producto:
        print(f"   - IP {ip.codigo_ip} (IR {ip.codigo_ir}) "
              f"metas={ip.metas_por_anio} pct={ip.escala_pct}")
